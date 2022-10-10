import sys , os
from openpyxl import workbook
from openpyxl.reader.excel import load_workbook
folder_arg = sys.argv[1]
search_text_arg = sys.argv[2]

def convert_to_unicode(text):
    text = str(text)
    Intermediate_code = (text).encode("utf-16")
    unicode_text = (Intermediate_code).decode("utf-16")
    return unicode_text

def find_text_in_multiple_workbooks(folder,search_text):
    search_text=str(search_text).lower()
    rows_found = []
    cells_found = []
    for item in os.scandir(folder):
            if item.name.endswith('.xlsx') and (not item.name.startswith('~')) and item.is_file():
                wb_name = item.path
                wb = load_workbook(wb_name)
                for ws in wb:
                    for r, row in enumerate(ws.values):
                        for c, value in enumerate(row):
                            if search_text in str(value).lower():
                                rows_found.append(ws[r+1])
                                cells_found.append([wb_name,ws.title,r+1,c+1,value])
                            else:
                                continue
            else:
                continue
    return cells_found,rows_found

cells_found,rows_found = find_text_in_multiple_workbooks (folder_arg , search_text_arg)
if  cells_found != []:
    for row in cells_found:
            print(convert_to_unicode(row),"\n")
    for row in rows_found:
        p = []
        for cell in row:
            q = convert_to_unicode(cell.value)
            p.append(q)
        print(p,"\n")
else:
    print("Nothing Found")